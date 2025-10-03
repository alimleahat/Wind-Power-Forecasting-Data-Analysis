import csv
from openpyxl import load_workbook  # to read Excel files

def airports(path="Airports.csv file location"):
    #airport data file
    rows = csv.reader(open(path))
    next(rows)  # skip the header row
    out = []  # create empty list for results
    for city, lat, lon in rows:
        # add each airport to our list with city name and coordinates
        out.append([city, float(lat), float(lon)])
    return out

def airlines(path="Airlines.csv file location"):
    # airlines data file
    rows = csv.reader(open(path))
    next(rows)  # skip the header row
    out = []  # create empty list for results
    for name, spd, ff, pf in rows:
        # add each airline with name, speed, fuel factor and price factor
        out.append([name, float(spd), float(ff), float(pf)])
    return out

def connections(path="Connections.xlsx file location"):
    # load the Excel file with connection data
    wb = load_workbook(path, data_only=True)
    out = {}  # create empty dictionary for results
    
    # go through each sheet in the Excel file
    for sh in wb.worksheets:
        # get the airport name from sheet title
        name = sh.title.replace("_connections", "")
        mat = []  # create empty list for this airport's connections
        
        # read all rows starting from row 2
        for row in sh.iter_rows(min_row=2, values_only=True):
            # add row to our matrix (skip first column)
            mat.append(list(row)[1:])
            
        # store the connection matrix for this airport
        out[name] = mat
    return out
